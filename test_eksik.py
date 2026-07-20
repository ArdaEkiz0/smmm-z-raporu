import os
import sys
import tempfile
import json
from unittest.mock import patch, MagicMock

import pytest
from PIL import Image


# ── OCR fonksiyonları ──

def _import_ocr():
    if "pytesseract" not in sys.modules:
        sys.modules["pytesseract"] = MagicMock()
    if "cv2" not in sys.modules:
        sys.modules["cv2"] = MagicMock()
    if "numpy" not in sys.modules:
        sys.modules["numpy"] = MagicMock()
    if "pyzbar" not in sys.modules:
        sys.modules["pyzbar"] = MagicMock()
        sys.modules["pyzbar.pyzbar"] = MagicMock()
    from ocr import ocr_guvenli, duzeltme_sozlugu, ogrenilen_sozluk
    return ocr_guvenli, duzeltme_sozlugu, ogrenilen_sozluk


def test_ocr_guvenli_bos_gorsel():
    ocr_guvenli, _, _ = _import_ocr()
    img = Image.new("L", (100, 30), color=255)
    result = ocr_guvenli(img)
    assert isinstance(result, str)


def test_ocr_guvenli_none():
    ocr_guvenli, _, _ = _import_ocr()
    result = ocr_guvenli(None)
    assert result == ""


def test_duzeltme_sozlugu_yukle():
    _, duzeltme_sozlugu, _ = _import_ocr()
    sonuc = duzeltme_sozlugu()
    assert isinstance(sonuc, dict)
    assert "BRÜT" in sonuc.values() or "BRÜT" in sonuc


def test_ogrenilen_sozluk_yukle():
    _, _, ogrenilen_sozluk = _import_ocr()
    sonuc = ogrenilen_sozluk()
    assert isinstance(sonuc, dict)


# ── Veritabanı fonksiyonları ──

def test_fis_kayit_bos():
    from veritabani import fis_kayit_bul
    with patch("veritabani.gecmis_listele", return_value=[]):
        fis, kayit = fis_kayit_bul("01.01.2026", "1")
        assert fis is None
        assert kayit is None


def test_fis_sil_olmayan_kayit():
    from veritabani import fis_sil
    with patch("veritabani.gecmis_listele", return_value=[]):
        sonuc = fis_sil("01.01.2026", "99999")
        assert sonuc is False


# ── Luca fonksiyonları ──

def _import_luca():
    for mod in ("numpy", "cv2"):
        sys.modules.pop(mod, None)
    if "pandas" not in sys.modules:
        sys.modules["pandas"] = MagicMock()
    from luca import generate_mukellef_rapor, generate_basit_usul_excel, HESAP_PLANLARI
    return generate_mukellef_rapor, generate_basit_usul_excel, HESAP_PLANLARI


def test_generate_mukellef_rapor():
    generate_mukellef_rapor, _, _ = _import_luca()
    fisler = [
        {"tarih": "01.07.2026", "z_no": "1", "brut": 1000, "net_toplam": 900,
         "nakit": 500, "kredi_karti": 400, "iadeler": 0, "banka_adi": "Ziraat"},
    ]
    html = generate_mukellef_rapor(fisler, "Test Mükellef <script>alert(1)</script>")
    assert "<!DOCTYPE html>" in html
    assert "Test Mükellef &lt;script&gt;" in html
    assert "<script>" not in html


def test_generate_basit_usul_excel():
    _, generate_basit_usul_excel, _ = _import_luca()
    results = [
        {"tarih": "01.07.2026", "z_no": "1", "brut": 500, "net_toplam": 450,
         "kredi_karti": 200, "toplam_tahsilat": 500,
         "urunler": [{"urun": "Ekmek", "miktar": 10, "tutar": 500, "oran": 1}]},
    ]
    muk_bilgi = {"adi": "Test Firma", "vergi_no": "1234567890", "vd": "Kadıköy"}
    excel_bytes = generate_basit_usul_excel(results, muk_bilgi)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0


def test_hesap_planlari_tanimli():
    _, _, HESAP_PLANLARI = _import_luca()
    assert isinstance(HESAP_PLANLARI, dict)
    for plan in ["LUCA", "Logo", "Netsis"]:
        assert plan in HESAP_PLANLARI
        for key in ["kredi_karti", "nakit", "satis_1"]:
            assert key in HESAP_PLANLARI[plan]


# ── Utils fonksiyonları ──

def test_dosya_yaz_oku():
    from utils import dosya_yaz, dosya_oku
    with tempfile.NamedTemporaryFile(mode="w", suffix=".json", delete=False) as f:
        path = f.name
    try:
        veri = {"test": "değer", "sayi": 42}
        dosya_yaz(path, veri)
        sonuc = dosya_oku(path)
        assert sonuc == veri
    finally:
        os.unlink(path)


def test_parse_tutar_binlik():
    from utils import parse_tutar
    assert parse_tutar("1.234,56") == 1234.56
    assert parse_tutar("10.000,00") == 10000.00
    assert parse_tutar("1.234.567,89") == 1234567.89


def test_parse_tutar_bos():
    from utils import parse_tutar
    assert parse_tutar("") == 0.0
    assert parse_tutar(None) == 0.0
    assert parse_tutar("abc") == 0.0


def test_levenshtein_bos():
    from utils import levenshtein
    assert levenshtein("", "") == 0
    assert levenshtein("abc", "abc") == 0


def test_levenshtein_farkli():
    from utils import levenshtein
    assert levenshtein("kitten", "sitting") == 3
    assert levenshtein("", "abc") == 3


# ── Yeni: Türkçe karakter düzeltme genişletmeleri ──


def test_duzeltme_kređi():
    _, duzeltme_sozlugu, _ = _import_ocr()
    from ocr import duzeltme_uygula
    assert "KREDİ" in duzeltme_uygula("KREĐI KARTI")
    assert "KREDİ" in duzeltme_uygula("KRDI KART")


def test_duzeltme_brut():
    _, duzeltme_sozlugu, _ = _import_ocr()
    from ocr import duzeltme_uygula, ocr_duzelt
    assert "BRÜT" in duzeltme_uygula("BURUT")
    assert "BRÜT" in duzeltme_uygula("BRUT")
    assert "BRÜT" in duzeltme_uygula("8RUT")
    assert "BRÜT" in ocr_duzelt("B R U T")


def test_duzeltme_nakit():
    _, _, _ = _import_ocr()
    from ocr import duzeltme_uygula
    assert "NAKİT" in duzeltme_uygula("NAKIT")


def test_duzeltme_tarih():
    _, _, _ = _import_ocr()
    from ocr import duzeltme_uygula
    result = duzeltme_uygula("TARIH")
    assert "TARİH" in result


def test_duzeltme_iade():
    _, _, _ = _import_ocr()
    from ocr import duzeltme_uygula
    assert "İADE" in duzeltme_uygula("IADE")
    assert "İADE" in duzeltme_uygula("1ADE")


def test_duzeltme_fis():
    _, _, _ = _import_ocr()
    from ocr import duzeltme_uygula
    assert "FİŞ" in duzeltme_uygula("FIS")


def test_duzeltme_kumulatif():
    _, _, _ = _import_ocr()
    from ocr import duzeltme_uygula
    assert "KÜMÜLATİF" in duzeltme_uygula("KUMULATIF")


def test_duzeltme_karakter_korunur():
    _, _, _ = _import_ocr()
    from ocr import duzeltme_uygula
    text = "NAKİT TOPLAM 100,00"
    assert duzeltme_uygula(text) == text


def test_duzeltme_coklu_yanlis():
    _, _, _ = _import_ocr()
    from ocr import duzeltme_uygula
    text = "BURUT NAKIT KREĐI KARTI ILE"
    result = duzeltme_uygula(text)
    assert "BRÜT" in result
    assert "NAKİT" in result
    assert "KREDİ" in result


# ── Yeni: Görüntü ön işleme ──


def test_gorsel_hazirla_default():
    _, _, _ = _import_ocr()
    from ocr import gorsel_hazirla
    img = Image.new("L", (800, 400), color=200)
    result = gorsel_hazirla(img, mode="default")
    assert result is not None
    assert result.mode == "L"
    assert result.size[0] >= 1200


def test_gorsel_hazirla_sayisal():
    _, _, _ = _import_ocr()
    from ocr import gorsel_hazirla
    img = Image.new("L", (800, 400), color=200)
    result = gorsel_hazirla(img, mode="sayisal")
    assert result is not None
    assert result.size[0] >= 1500


def test_gorsel_hazirla_kalin():
    _, _, _ = _import_ocr()
    from ocr import gorsel_hazirla
    img = Image.new("L", (800, 400), color=200)
    result = gorsel_hazirla(img, mode="kalin")
    assert result is not None
    assert result.size[0] >= 1200


def test_gorsel_hazirla_buyuk_gorsel():
    _, _, _ = _import_ocr()
    from ocr import gorsel_hazirla
    img = Image.new("L", (2000, 1000), color=128)
    result = gorsel_hazirla(img, mode="default")
    assert result is not None
    assert result.size == (2000, 1000) or result.size[0] < 2000


def test_clahe_yardimci():
    _, _, _ = _import_ocr()
    from ocr import _clahe
    img = Image.new("L", (200, 200), color=128)
    result = _clahe(img)
    assert result is not None
    assert result.size == (200, 200)


def test_bilateral_denoise_yardimci():
    _, _, _ = _import_ocr()
    from ocr import _bilateral_denoise
    img = Image.new("L", (200, 200), color=128)
    result = _bilateral_denoise(img)
    assert result is not None
    assert result.size == (200, 200)


def test_deskew_dusuk_egim():
    _, _, _ = _import_ocr()
    from ocr import _deskew
    img = Image.new("L", (200, 200), color=128)
    result = _deskew(img)
    assert result is not None


def test_remove_border():
    _, _, _ = _import_ocr()
    from ocr import _remove_border
    img = Image.new("L", (300, 300), color=255)
    pixels = img.load()
    for x in range(50, 250):
        pixels[x, 50] = 0
        pixels[x, 250] = 0
    for y in range(50, 250):
        pixels[50, y] = 0
        pixels[250, y] = 0
    result = _remove_border(img)
    assert result is not None


# ── Yeni: OCR image pipeline ──


def test_ocr_image_bos_gorsel():
    _, _, _ = _import_ocr()
    from ocr import ocr_image
    img = Image.new("L", (100, 30), color=255)
    with patch("ocr.ocr_guvenli", return_value="NAKİT 100,00\nTOPLAM 100,00"):
        with patch("ocr.gorsel_hazirla", return_value=img):
            with patch("ocr._ocr_hazirla_otsu", return_value=img):
                with patch("ocr.ocr_skorla", return_value=10):
                    result = ocr_image(img)
    assert isinstance(result, str)


def test_ocr_image_tekrarli_satirlar_temizleme():
    _, _, _ = _import_ocr()
    from ocr import ocr_image
    img = Image.new("L", (100, 30), color=255)
    dup_text = "NAKİT 100,00\nNAKİT 100,00\nTOPLAM"
    with patch("ocr.ocr_guvenli", return_value=dup_text):
        with patch("ocr.gorsel_hazirla", return_value=img):
            with patch("ocr._ocr_hazirla_otsu", return_value=img):
                with patch("ocr.ocr_skorla", return_value=100):
                    result = ocr_image(img)
    assert "NAKİT 100,00" in result


def test_ocr_image_birden_fazla_psm():
    _, _, _ = _import_ocr()
    from ocr import ocr_image
    img = Image.new("L", (100, 30), color=255)

    call_count = {"n": 0}

    def fake_ocr(img, psm=6, config_extra=""):
        call_count["n"] += 1
        if psm == 3:
            return "BRÜT 500,00"
        if psm == 4:
            return "NAKİT 200,00"
        return "TOPLAM 700,00"

    with patch("ocr.ocr_guvenli", side_effect=fake_ocr):
        with patch("ocr.gorsel_hazirla", return_value=img):
            with patch("ocr._ocr_hazirla_otsu", return_value=img):
                with patch("ocr.ocr_skorla", return_value=10):
                    result = ocr_image(img)

    assert call_count["n"] >= 3


def test_ocr_image_buyuk_olcekleme():
    _, _, _ = _import_ocr()
    from ocr import ocr_image
    img = Image.new("L", (100, 30), color=255)
    with patch("ocr.ocr_guvenli", return_value="TOPLAM 100,00"):
        with patch("ocr.gorsel_hazirla", return_value=img) as mh:
            with patch("ocr._ocr_hazirla_otsu", return_value=img):
                with patch("ocr.ocr_skorla", return_value=50):
                    ocr_image(img)
    assert mh.called


def test_ocr_image_donus_180():
    _, _, _ = _import_ocr()
    from ocr import ocr_image
    img = Image.new("L", (100, 30), color=255)
    with patch("ocr.ocr_guvenli", return_value="BURUT 100"):
        with patch("ocr.gorsel_hazirla", return_value=img):
            with patch("ocr._ocr_hazirla_otsu", return_value=img):
                with patch("ocr.ocr_skorla", return_value=10):
                    result = ocr_image(img)
    assert isinstance(result, str)


# ── Yeni: LUCA Export doğrulama testleri ──


def _luca_balance(rows):
    b = sum(r.get("Borç", 0) or 0 for r in rows)
    a = sum(r.get("Alacak", 0) or 0 for r in rows)
    return b, a


def test_luca_data_to_luca_rows_temel():
    from luca import data_to_luca_rows, varsayilan_kodlar
    data = {
        "tarih": "01.07.2026", "z_no": "100", "belge_no": "100",
        "nakit": 5000, "kredi_karti": 0, "yemek_ceki": 0,
        "toplam_tahsilat": 5000, "iadeler": 0, "net_toplam": 5000,
        "brut": 5000, "urunler": [], "banka_adi": None, "firma_adi": None,
    }
    rows = data_to_luca_rows(data, varsayilan_kodlar(), 1, [])
    assert len(rows) > 0
    for r in rows:
        assert r["İŞLEM"] == "1"
        assert r["BELGE TÜRÜ"] == "Z Raporu"
    b, a = _luca_balance(rows)
    assert abs(b - a) < 0.01, f"Dengesiz: B={b:.2f} A={a:.2f}"


def test_luca_data_to_luca_rows_iade_kk():
    from luca import data_to_luca_rows, varsayilan_kodlar
    data = {
        "tarih": "01.07.2026", "z_no": "101", "belge_no": "101",
        "nakit": 0, "kredi_karti": 5000, "yemek_ceki": 0,
        "toplam_tahsilat": 5000, "iadeler": 2000, "net_toplam": 3000,
        "brut": 5000, "kdv_kalemleri": [{"oran": 20, "matrah": 4166.67, "kdv_tutari": 833.33}],
        "urunler": [], "banka_adi": None, "firma_adi": None,
    }
    rows = data_to_luca_rows(data, varsayilan_kodlar(), 1, [])
    kk_rows = [r for r in rows if "KK" in r.get("AÇIKLAMA", "")]
    assert len(kk_rows) == 1
    assert abs(kk_rows[0]["Borç"] - 3000) < 0.01
    b, a = _luca_balance(rows)
    assert abs(b - a) < 0.01


def test_luca_tum_planlar_dengeli():
    from luca import data_to_luca_rows, HESAP_PLANLARI
    data = {
        "tarih": "01.07.2026", "z_no": "1", "nakit": 3000, "kredi_karti": 2000,
        "net_toplam": 5000, "brut": 5000, "toplam_tahsilat": 5000,
        "iadeler": 0, "urunler": [], "kdv_kalemleri": [],
        "banka_adi": "", "firma_adi": "",
    }
    for plan_adi, plan_kodlar in HESAP_PLANLARI.items():
        rows = data_to_luca_rows(data, plan_kodlar, 1, [])
        b, a = _luca_balance(rows)
        assert abs(b - a) < 0.01, f"{plan_adi} dengesiz: B={b:.2f} A={a:.2f}"


def test_luca_excel_bytes():
    from luca import data_to_luca_rows, generate_excel, varsayilan_kodlar
    data = {
        "tarih": "01.07.2026", "z_no": "200", "belge_no": "200",
        "nakit": 5000, "kredi_karti": 0, "net_toplam": 5000, "brut": 5000,
        "toplam_tahsilat": 5000, "iadeler": 0, "urunler": [],
        "banka_adi": "", "firma_adi": "",
    }
    rows = data_to_luca_rows(data, varsayilan_kodlar(), 1, [])
    excel = generate_excel(rows)
    assert isinstance(excel, bytes)
    assert len(excel) > 100


def test_luca_urun_kodu_eslesmesi():
    from luca import data_to_luca_rows, varsayilan_kodlar, urun_kodlari_varsayilan
    data = {
        "tarih": "01.07.2026", "z_no": "102", "belge_no": "102",
        "nakit": 1200, "kredi_karti": 0, "yemek_ceki": 0,
        "toplam_tahsilat": 1200, "iadeler": 0, "net_toplam": 1200,
        "brut": 1200,
        "urunler": [{"urun": "EKMEK", "oran": 20, "miktar": 10, "tutar": 1200}],
        "banka_adi": None, "firma_adi": None,
    }
    urun_kodlari = urun_kodlari_varsayilan()
    rows = data_to_luca_rows(data, varsayilan_kodlar(), 1, urun_kodlari)
    ekmek_rows = [r for r in rows if r["Hesap Kodu"] == "600.06"]
    assert len(ekmek_rows) == 1


def test_luca_basit_usul_kolon_sayisi():
    from luca import generate_basit_usul_excel
    from config import BASIT_USUL_KOLONLAR
    results = [
        {"tarih": "01.07.2026", "z_no": "1", "brut": 1000, "net_toplam": 900,
         "nakit": 500, "kredi_karti": 400, "toplam_tahsilat": 1000,
         "iadeler": 0, "banka_adi": "Ziraat", "firma_adi": "Test",
         "urunler": []},
    ]
    excel = generate_basit_usul_excel(results, {"adi": "Test"})
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(excel))
    ws = wb.active
    assert ws.max_column == len(BASIT_USUL_KOLONLAR), \
        f"Beklenen={len(BASIT_USUL_KOLONLAR)} sutun, var={ws.max_column}"


def test_luca_senaryo_dengesi():
    from luca import data_to_luca_rows, varsayilan_kodlar
    senaryolar = [
        (5000, 0, 0, 0),
        (3000, 7000, 0, 0),
        (2000, 5000, 3000, 0),
        (5000, 5000, 0, 500),
        (5000, 5000, 0, 3000),
        (1000, 2000, 500, 3500),
    ]
    kodlar = varsayilan_kodlar()
    for nakit, kk, yemek, iade in senaryolar:
        brut = nakit + kk + yemek
        matrah = round(brut / 1.2, 2) if brut > 0 else 0
        kdv = round(brut - matrah, 2) if brut > 0 else 0
        data = {
            "tarih": "01.07.2026", "z_no": "999", "belge_no": "999",
            "nakit": nakit, "kredi_karti": kk, "yemek_ceki": yemek,
            "toplam_tahsilat": brut, "iadeler": iade,
            "net_toplam": brut - iade, "brut": brut,
            "kdv_kalemleri": [{"oran": 20, "matrah": matrah, "kdv_tutari": kdv}] if brut > 0 else [],
            "urunler": [], "banka_adi": None, "firma_adi": None,
        }
        rows = data_to_luca_rows(data, kodlar, 1, [])
        b, a = _luca_balance(rows)
        assert abs(b - a) < 0.01, f"nakit={nakit} kk={kk} yemek={yemek} iade={iade}: B={b:.2f} A={a:.2f}"
