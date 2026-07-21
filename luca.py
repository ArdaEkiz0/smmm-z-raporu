import io
import html as html_mod
from datetime import datetime

from config import BASIT_USUL_KOLONLAR, SABLON_FILE, URUN_KODLARI_FILE
from utils import dosya_oku, log, parse_tutar


def varsayilan_kodlar():
    return {
        "kredi_karti": "108.01", "nakit": "100.01", "yemek_ceki": "108.03",
        "satis_1": "600.01", "satis_10": "600.05",
        "satis_20": "600.04",
        "kdv_1": "391.01", "kdv_10": "391.05",
        "kdv_20": "391.04", "iadeler": "610.01",
    }


HESAP_PLANLARI = {
    "LUCA": varsayilan_kodlar(),
    "Logo": {
        "kredi_karti": "100.02", "nakit": "100.01", "yemek_ceki": "100.03",
        "satis_1": "601.01", "satis_10": "601.02",
        "satis_20": "601.03",
        "kdv_1": "391.01", "kdv_10": "391.02",
        "kdv_20": "391.03", "iadeler": "602.01",
    },
    "Netsis": {
        "kredi_karti": "120.01", "nakit": "100.01", "yemek_ceki": "120.02",
        "satis_1": "600.01", "satis_10": "600.02",
        "satis_20": "600.03",
        "kdv_1": "391.01", "kdv_10": "391.02",
        "kdv_20": "391.03", "iadeler": "610.01",
    },
}


def urun_kodlari_varsayilan():
    return [
        {"pattern": "EKMEK", "hesap_kodu": "600.06", "aciklama": "Ekmek Satışı", "kdv_orani": 1},
        {"pattern": "SİGARA", "hesap_kodu": "600.07", "aciklama": "Sigara Satışı", "kdv_orani": 0},
        {"pattern": "SIGARA", "hesap_kodu": "600.07", "aciklama": "Sigara Satışı", "kdv_orani": 0},
        {"pattern": "SÜT", "hesap_kodu": "600.08", "aciklama": "Süt ve Süt Ürünleri", "kdv_orani": 10},
    ]


def urun_kodlari_yukle():
    return dosya_oku(URUN_KODLARI_FILE, urun_kodlari_varsayilan())


def urun_kodlari_kaydet(kodlar):
    dosya_yaz(URUN_KODLARI_FILE, kodlar)


def urun_kodu_bul(urun_kodlari, urun_adi):
    if not urun_adi or not urun_kodlari:
        return None
    ua = urun_adi.upper().strip()
    import re
    for uk in urun_kodlari:
        pat = uk.get("pattern", "")
        if pat and re.search(re.escape(pat), ua, re.IGNORECASE):
            return uk
    return None


def _iade_dagit(iade, nakit, kk, yemek):
    """İade tutarını kaynaklara (nakit, KK, yemek) sırayla dağıt.
    Döndürür: (iade_nkt_k, iade_kk_k, iade_yem_k, yeni_nakit, yeni_kk, yeni_yemek)."""
    if iade <= 0:
        return 0, 0, 0, nakit, kk, yemek
    iade_kk_k = round(min(kk, iade), 2) if kk > 0 else 0
    iade_nkt_k = round(min(nakit, iade - iade_kk_k), 2) if nakit > 0 and iade > iade_kk_k else 0
    iade_yem_k = round(iade - iade_kk_k - iade_nkt_k, 2)
    yeni_nakit = round(nakit - iade_nkt_k, 2) if iade_nkt_k > 0 else nakit
    yeni_kk = round(kk - iade_kk_k, 2) if iade_kk_k > 0 else kk
    yeni_yemek = round(yemek - iade_yem_k, 2) if iade_yem_k > 0 else yemek
    return iade_nkt_k, iade_kk_k, iade_yem_k, yeni_nakit, yeni_kk, yeni_yemek


def data_to_luca_rows(data, hesap_kodlari, fis_no=1, urun_kodlari=None):
    rows = []

    def satir(hesap_kodu, aciklama, borc, alacak):
        return {
            "İŞLEM": "1", "KATEGORİ": "Defter Fişleri", "BELGE TÜRÜ": "Z Raporu",
            "BELGE TARİHİ": data.get("tarih", ""), "FİŞ TARİHİ": data.get("tarih", ""),
            "FİŞ NO": str(fis_no), "BELGE NO": data.get("z_no", ""),
            "MÜKELLEF/ALICI TC KİMLİK NO": "", "BAĞLI OLDUĞU VERGİ DAİRESİ": "",
            "MÜKELLEF/ALICI ÜNVAN": data.get("firma_adi", "") or data.get("mukellef_adi", ""),
            "MÜKELLEF/ALICI SOYADI": "", "ADRES": "", "PLAKA NO": "", "KİLOMETRE": "",
            "CİNSİ": "", "GİDER TÜRÜ": "", "KDV Oranları GİRİŞ": "", "KDV Oranları ÇIKIŞ": "",
            "STOK KODU": "", "MALZEME/HİZMET ADI": aciklama,
            "MİKTAR": "", "BİRİM FİYAT": "", "TUTAR": "", "TUTAR (TL)": "",
            "KDV ORANI (%)": "", "İSKONTO ORANI (%)": "", "İSKONTO TUTARI (TL)": "",
            "VERGİLER DAHİL TOPLAM": "", "KDV TUTARI": "", "GENEL TOPLAM": "",
            "KREDİ KARTI İLE TAHSİLAT": "", "TAHSİL EDİLEN": "",
            "KALEM SAYISI": "", "SIRA NO": "", "ÖZEL KOD": "",
            "AÇIKLAMA": aciklama,
            "Hesap Kodu": hesap_kodu, "Borç": borc, "Alacak": alacak,
        }

    net_toplam = data.get("net_toplam", 0) or 0
    nakit = data.get("nakit", 0) or 0
    kk = data.get("kredi_karti", 0) or 0
    yemek = data.get("yemek_ceki", 0) or 0
    iade = data.get("iadeler", 0) or 0
    musteri = data.get("firma_adi", "") or data.get("mukellef_adi", "") or "Müşteri"
    brut = data.get("brut", 0) or 0

    # Z raporundan KDV
    kdv_kalemleri = data.get("kdv_kalemleri", [])
    urunler = data.get("urunler", [])
    if kdv_kalemleri:
        z_toplam_kdv = sum(k.get("kdv_tutari", 0) or 0 for k in kdv_kalemleri)
        if z_toplam_kdv <= 0:
            for k in kdv_kalemleri:
                matrah = k.get("matrah", 0) or 0
                oran = k.get("oran", 0) or 0
                if matrah > 0 and oran > 0:
                    z_toplam_kdv += round(matrah * oran / (100 + oran), 2)
    elif urunler:
        z_toplam_kdv = 0
        for u in urunler:
            oran = u.get("oran", 0) or 0
            tutar = u.get("tutar", 0) or 0
            if oran > 0:
                z_toplam_kdv += round(tutar - tutar / (1 + oran / 100), 2)
    elif brut > 0 and net_toplam > 0 and net_toplam < brut:
        z_toplam_kdv = round(brut - net_toplam, 2)
    else:
        z_toplam_kdv = 0

    tahsilat_toplam = nakit + kk + yemek

    # İade varsa, tahsilat kaynaklarına dağıt (önce KK, sonra nakit, sonra yemek)
    # ve 610.01'de ayrı bir borç satırı olarak göster
    if iade > 0:
        iade_nkt_k, iade_kk_k, iade_yem_k, nakit, kk, yemek = _iade_dagit(iade, nakit, kk, yemek)

    def _tahsilat_ekle(rows):
        if iade > 0:
            rows.append(satir(hesap_kodlari.get("iadeler", "610.01"), f"İade - {musteri}", iade, 0))
        if nakit > 0:
            rows.append(satir(hesap_kodlari.get("nakit", "100.01"), f"Nakit Tahsilat - {musteri}", nakit, 0))
        if kk > 0:
            rows.append(satir(hesap_kodlari.get("kredi_karti", "108.01"), f"KK Tahsilat - {musteri}", kk, 0))
        if yemek > 0:
            rows.append(satir(hesap_kodlari.get("yemek_ceki", "108.03"), f"Yemek Çeki - {musteri}", yemek, 0))

    # Ürün bazlı satışlar - Z raporu tutarlarini dogrudan kullan
    if urunler:
        toplam_urun_tutari = sum((u.get("tutar", 0) or 0) for u in urunler)

        # Tum urunlerin oran=0 ama Z raporunda KDV varsa, KDV'yi orantili olarak matrahlardan cikar
        tum_oran_sifir = all((u.get("oran", 0) or 0) == 0 for u in urunler)
        if tum_oran_sifir and z_toplam_kdv > 0 and toplam_urun_tutari > 0:
            toplam_matrah = round(toplam_urun_tutari - z_toplam_kdv, 2)
            for u in urunler:
                urun_adi = u.get("urun", "Ürün")
                tutar = u.get("tutar", 0) or 0
                oran_pay = tutar / toplam_urun_tutari if toplam_urun_tutari > 0 else 0
                matrah = round(toplam_matrah * oran_pay, 2)
                satis_kod = "satis_0"
                hesap_kodu = hesap_kodlari.get(satis_kod, hesap_kodlari.get("satis_20", "600.04"))
                satis_idx = len(rows)
                rows.append(satir(hesap_kodu, f"{urun_adi} - {musteri}", 0, matrah))
                if urun_kodlari:
                    eslesme = urun_kodu_bul(urun_kodlari, urun_adi)
                    if eslesme:
                        rows[satis_idx]["Hesap Kodu"] = eslesme.get("hesap_kodu", "")
            # Z raporundaki KDV'leri ekle
            for kv in data.get("kdv_kalemleri", []):
                kv_oran = kv.get("oran", 0) or 0
                kv_tutar = kv.get("kdv_tutari", 0) or 0
                if kv_tutar > 0:
                    kdv_kod = "kdv_" + str(kv_oran)
                    kdv_hk = hesap_kodlari.get(kdv_kod, hesap_kodlari.get("kdv_1", "391.01"))
                    rows.append(satir(kdv_hk, f"KDV %{kv_oran} - {musteri}", 0, kv_tutar))
        else:
            for u in urunler:
                urun_adi = u.get("urun", "Ürün")
                tutar = u.get("tutar", 0) or 0
                kdv_orani = u.get("oran", 0) or 0
                # Z raporundaki tutar KDV DAHIL (brut) degerdir, icinden KDV'yi cikar
                if kdv_orani > 0:
                    matrah = round(tutar / (1 + kdv_orani / 100), 2)
                    kdv = round(tutar - matrah, 2)
                else:
                    matrah = tutar
                    kdv = 0
                satis_kod = "satis_" + str(kdv_orani)
                hesap_kodu = hesap_kodlari.get(satis_kod, hesap_kodlari.get("satis_20", "600.04"))
                satis_idx = len(rows)
                rows.append(satir(hesap_kodu, f"{urun_adi} - {musteri}", 0, matrah))
                if kdv > 0:
                    kdv_kod = "kdv_" + str(kdv_orani)
                    kdv_hk = hesap_kodlari.get(kdv_kod, hesap_kodlari.get("kdv_20", "391.04"))
                    rows.append(satir(kdv_hk, f"KDV %{kdv_orani} - {urun_adi} - {musteri}", 0, kdv))

                if urun_kodlari:
                    eslesme = urun_kodu_bul(urun_kodlari, urun_adi)
                    if eslesme:
                        urun_kodu = eslesme.get("hesap_kodu", "")
                        rows[satis_idx]["Hesap Kodu"] = urun_kodu

        # Dengeleme kontrolu: Urunlerden hesaplanan KDV + Matrah, Z raporu degerleriyle eslesmeli
        # Brüt/Net referans alinarak dengelenir
        urun_toplam_kdv = sum((r.get("Alacak", 0) or 0) for r in rows if "KDV" in r.get("AÇIKLAMA", ""))
        urun_toplam_matrah = sum((r.get("Alacak", 0) or 0) for r in rows if "KDV" not in r.get("AÇIKLAMA", ""))
        urun_toplam_alacak = urun_toplam_kdv + urun_toplam_matrah

        # Hedef: toplam_alacak = brut (urunler KDV dahil toplami)
        # brut Z raporundan gelir; eger urunlerle uyusmuyorsa, son matrah satirina fark ekle
        beklenen_toplam = brut if brut > 0 else toplam_urun_tutari
        toplam_fark = round(beklenen_toplam - urun_toplam_alacak, 2)
        if abs(toplam_fark) > 0.01 and rows:
            # Farki son matrah satirina uygula (KDV degistirilmez, brüt referans)
            matrah_satirlari = [r for r in rows if "KDV" not in r.get("AÇIKLAMA", "")]
            if matrah_satirlari:
                son_matrah = matrah_satirlari[-1]
                eski_matrah = son_matrah.get("Alacak", 0) or 0
                son_matrah["Alacak"] = round(eski_matrah + toplam_fark, 2)
            else:
                # Matrah satiri yoksa, son KDV satırına farkı ekle
                kdv_satirlari = [r for r in rows if "KDV" in r.get("AÇIKLAMA", "")]
                if kdv_satirlari:
                    son_kdv = kdv_satirlari[-1]
                    eski_kdv = son_kdv.get("Alacak", 0) or 0
                    son_kdv["Alacak"] = round(eski_kdv + toplam_fark, 2)

        # KDV satiri hic yoksa, hesaplanan KDV'yi tek satir olarak ekle
        kdv_satirlari = [r for r in rows if "KDV" in r.get("AÇIKLAMA", "")]
        if urun_toplam_kdv > 0 and not kdv_satirlari:
            kdv_kalemleri_z = data.get("kdv_kalemleri", [])
            if kdv_kalemleri_z:
                for kv in kdv_kalemleri_z:
                    kv_oran = kv.get("oran", 0) or 0
                    kv_tutar = kv.get("kdv_tutari", 0) or 0
                    if kv_tutar > 0:
                        kdv_kod = "kdv_" + str(kv_oran)
                        kdv_hk = hesap_kodlari.get(kdv_kod, hesap_kodlari.get("kdv_20", "391.04"))
                        rows.append(satir(kdv_hk, f"KDV %{kv_oran} - {musteri}", 0, kv_tutar))
            else:
                rows.append(satir(hesap_kodlari.get("kdv_1", "391.01"), f"KDV - {musteri}", 0, urun_toplam_kdv))

        _tahsilat_ekle(rows)
        return rows

    # Ürün yok - klasik toplu muhasebe
    kdv_kalemleri = data.get("kdv_kalemleri", [])
    if kdv_kalemleri:
        for kv in kdv_kalemleri:
            oran = kv.get("oran", 0)
            matrah = kv.get("matrah", 0) or 0
            kdv_t = kv.get("kdv_tutari", 0) or 0
            satis_key = "satis_" + str(oran)
            if matrah > 0:
                rows.append(satir(hesap_kodlari.get(satis_key, "600.04"), f"Satış %{oran} - {musteri}", 0, matrah))
                kdv_key = "kdv_" + str(oran)
                rows.append(satir(hesap_kodlari.get(kdv_key, "391.04"), f"KDV %{oran} - {musteri}", 0, kdv_t))
        _tahsilat_ekle(rows)
        return rows

    # En basit: tek satır satış - Z raporu değerlerini doğrudan kullan
    if net_toplam <= 0 and nakit <= 0 and kk <= 0 and yemek <= 0:
        return rows

    rows.append(satir("600.04", f"Z Raporu Satış - {musteri}", 0, net_toplam))
    if z_toplam_kdv > 0:
        rows.append(satir("391.04", f"KDV - {musteri}", 0, z_toplam_kdv))
    _tahsilat_ekle(rows)
    return rows


def hesapla_luca_rows(results, hesap_kodlari, urun_kodlari):
    all_luca_rows = []
    fc = 1
    for r in results:
        if "error" not in r:
            rows = data_to_luca_rows(r, hesap_kodlari, fc, urun_kodlari)
            all_luca_rows.extend(rows)
            fc += 1
    return all_luca_rows


def generate_mukellef_rapor(fisler, mukellef_adi):
    nakit_toplam = sum(f.get("nakit", 0) or 0 for f in fisler)
    kk_toplam = sum(f.get("kredi_karti", 0) or 0 for f in fisler)
    iade_toplam = sum(f.get("iadeler", 0) or 0 for f in fisler)
    net_toplam = sum(f.get("net_toplam", 0) or 0 for f in fisler)
    brüt_toplam = sum(f.get("brut", 0) or 0 for f in fisler)

    satirlar = ""
    for f in fisler:
        satirlar += f"""<tr>
<td>{html_mod.escape(str(f.get('tarih','?')))}</td>
<td>{html_mod.escape(str(f.get('z_no','?')))}</td>
<td>{html_mod.escape(str(f.get('banka_adi','-')))}</td>
<td style='text-align:right'>{f.get('brut',0):,.2f}</td>
<td style='text-align:right'>{f.get('net_toplam',0):,.2f}</td>
<td style='text-align:right'>{f.get('nakit',0):,.2f}</td>
<td style='text-align:right'>{f.get('kredi_karti',0):,.2f}</td>
<td style='text-align:right'>{f.get('iadeler',0):,.2f}</td>
</tr>"""

    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    guvenli_adi = html_mod.escape(str(mukellef_adi))
    html_icerik = f"""<!DOCTYPE html><html><head><meta charset="utf-8">
<title>Mükellef Raporu - {guvenli_adi}</title>
<style>
body {{ font-family: Arial, sans-serif; margin: 20px; }}
h1 {{ text-align: center; color: #1a5276; }}
.meta {{ text-align: center; color: #666; }}
.ozet {{ display: flex; justify-content: space-around; margin: 20px 0; }}
.ozet > div {{ text-align: center; padding: 10px; background: #e8f6f3; border-radius: 8px; min-width: 120px; }}
.deger {{ font-size: 24px; font-weight: bold; color: #1a5276; }}
.etiket {{ font-size: 12px; color: #666; }}
table {{ width: 100%; border-collapse: collapse; margin: 15px 0; }}
th, td {{ border: 1px solid #ddd; padding: 6px; }}
th {{ background: #1a5276; color: white; }}
</style></head><body>
<h1>{guvenli_adi}</h1>
<p class="meta">{now} | {len(fisler)} Fiş</p>
<div class="ozet">
<div><div class="deger">{net_toplam:,.2f} ₺</div><div class="etiket">Net Toplam</div></div>
<div><div class="deger">{brüt_toplam:,.2f} ₺</div><div class="etiket">Brüt Toplam</div></div>
<div><div class="deger">{nakit_toplam:,.2f} ₺</div><div class="etiket">Nakit</div></div>
<div><div class="deger">{kk_toplam:,.2f} ₺</div><div class="etiket">Kredi Kartı</div></div>
<div><div class="deger">{iade_toplam:,.2f} ₺</div><div class="etiket">İade</div></div>
</div>
<table>
<tr><th>Tarih</th><th>Z No</th><th>Banka</th><th>Brüt</th><th>Net</th><th>Nakit</th><th>KK</th><th>İade</th></tr>
{satirlar}
<tr style="font-weight:bold;background:#e8f6f3">
<td colspan="3">TOPLAM</td>
<td style='text-align:right'>{brüt_toplam:,.2f}</td>
<td style='text-align:right'>{net_toplam:,.2f}</td>
<td style='text-align:right'>{nakit_toplam:,.2f}</td>
<td style='text-align:right'>{kk_toplam:,.2f}</td>
<td style='text-align:right'>{iade_toplam:,.2f}</td>
</tr>
</table>
<p style="text-align:center;color:#999;margin-top:30px">SMMM Z Raporu ve Fiş Yönetim Sistemi</p>
</body></html>"""
    return html_icerik


def generate_excel(data_rows, ozet_satiri=True):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    wb = Workbook()
    ws = wb.active
    ws.title = "Z Raporlari"

    if not data_rows:
        ws.cell(row=1, column=1, value="Veri Yok")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    kolonlar = list(data_rows[0].keys())
    for ek in ["Hesap Kodu", "Borç", "Alacak"]:
        if ek in kolonlar:
            kolonlar.remove(ek)

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    borc_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    alacak_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    toplam_font = Font(bold=True, size=10)
    toplam_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for ci, kolon_adi in enumerate(kolonlar, 1):
        cell = ws.cell(row=1, column=ci, value=kolon_adi)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    borc_kolon_idx = None
    alacak_kolon_idx = None
    for ci, kolon_adi in enumerate(kolonlar):
        if kolon_adi == "Borç":
            borc_kolon_idx = ci
        elif kolon_adi == "Alacak":
            alacak_kolon_idx = ci

    for ri, row in enumerate(data_rows, 2):
        for ci, kolon_adi in enumerate(kolonlar, 1):
            deger = row.get(kolon_adi, "")
            cell = ws.cell(row=ri, column=ci, value=deger)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if borc_kolon_idx is not None and ci - 1 == borc_kolon_idx and isinstance(deger, (int, float)) and deger > 0:
                cell.fill = borc_fill
            if alacak_kolon_idx is not None and ci - 1 == alacak_kolon_idx and isinstance(deger, (int, float)) and deger > 0:
                cell.fill = alacak_fill

    if ozet_satiri and data_rows:
        toplam_satir = len(data_rows) + 2
        ws.cell(row=toplam_satir, column=1, value="TOPLAM").font = toplam_font
        ws.cell(row=toplam_satir, column=1).fill = toplam_fill
        for ci in range(1, len(kolonlar) + 1):
            ws.cell(row=toplam_satir, column=ci).border = thin_border
            ws.cell(row=toplam_satir, column=ci).fill = toplam_fill
            ws.cell(row=toplam_satir, column=ci).font = toplam_font

        if borc_kolon_idx is not None:
            toplam_borc = sum((r.get("Borç", 0) or 0) for r in data_rows)
            cell = ws.cell(row=toplam_satir, column=borc_kolon_idx + 1, value=toplam_borc)
            cell.number_format = '#,##0.00'
        if alacak_kolon_idx is not None:
            toplam_alacak = sum((r.get("Alacak", 0) or 0) for r in data_rows)
            cell = ws.cell(row=toplam_satir, column=alacak_kolon_idx + 1, value=toplam_alacak)
            cell.number_format = '#,##0.00'

    for col_idx in range(1, len(kolonlar) + 1):
        max_len = 0
        for row_idx in range(1, min(len(data_rows) + 2, 50)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 45)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_excel_cached(data_rows_tuple):
    return generate_excel(list(data_rows_tuple))


def generate_basit_usul_excel(results, muk_bilgi, sablon_data=None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Serbest Meslek"

    if not results:
        ws.cell(row=1, column=1, value="Veri Yok")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    if sablon_data:
        try:
            from openpyxl import load_workbook
            sablon_io = io.BytesIO(sablon_data)
            wb2 = load_workbook(sablon_io)
            ws2 = wb2.active
            basliklar = []
            for col in ws2.iter_cols(1, ws2.max_column, max_row=1):
                basliklar.append(col[0].value or "")
            if basliklar:
                return sablon_data
        except Exception:
            log.warning("Şablon dosyası okunamadı", exc_info=True)

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1A5276", end_color="1A5276", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for ci, kolon_adi in enumerate(BASIT_USUL_KOLONLAR, 1):
        cell = ws.cell(row=1, column=ci, value=kolon_adi)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    tckn = (muk_bilgi or {}).get("vergi_no", "")
    vd = (muk_bilgi or {}).get("vd", "")
    unvan = (muk_bilgi or {}).get("adi", "")
    adres = (muk_bilgi or {}).get("notlar", "")

    def base_row():
        return [""] * len(BASIT_USUL_KOLONLAR)

    def yaz_satir(row, evrak_tarihi, evrak_no, ua, miktar, brut_tutar, oran, kk_ksm):
        row[0] = "1"
        row[1] = "Defter Fişleri"
        row[2] = "Z Raporu"
        row[3] = evrak_tarihi
        row[4] = evrak_tarihi
        row[6] = evrak_no
        row[7] = tckn
        row[8] = vd
        row[9] = unvan
        row[11] = adres
        row[19] = ua
        row[20] = miktar
        birim_f = round(brut_tutar / miktar, 2) if (isinstance(miktar, (int, float)) and miktar > 0) else ""
        row[21] = birim_f
        matrah = round(brut_tutar / (1 + oran / 100), 2) if oran > 0 else brut_tutar
        row[22] = matrah
        row[24] = oran if oran > 0 else ""
        kdv = round(brut_tutar - matrah, 2) if oran > 0 else 0
        row[28] = kdv
        row[29] = brut_tutar
        row[30] = kk_ksm

    ri = 2
    for r in results:
        if "error" in r:
            continue
        evrak_tarihi = r.get("tarih", "")
        evrak_no = r.get("z_no", "") or r.get("belge_no", "")
        kk_tutar = r.get("kredi_karti", 0) or 0
        toplam_tahsilat = r.get("toplam_tahsilat", 0) or r.get("brut", 0) or 1
        kk_orani = kk_tutar / max(toplam_tahsilat, 1)
        urunler = r.get("urunler", [])
        if not urunler:
            row = base_row()
            brut = r.get("brut", 0) or toplam_tahsilat or 0
            kk_ksm = round(brut * kk_orani, 2)
            yaz_satir(row, evrak_tarihi, evrak_no, "", "", brut, 0, kk_ksm)
            for ci, val in enumerate(row):
                ws.cell(row=ri, column=ci + 1, value=val).border = thin_border
            ri += 1
            continue
        for urun in urunler:
            row = base_row()
            ua = urun.get("urun", "")
            miktar = urun.get("miktar", 0) or 0
            brut_tutar = urun.get("tutar", 0) or 0
            oran = urun.get("oran", 0) or 0
            kk_ksm = round(brut_tutar * kk_orani, 2)
            yaz_satir(row, evrak_tarihi, evrak_no, ua, miktar, brut_tutar, oran, kk_ksm)
            for ci, val in enumerate(row):
                ws.cell(row=ri, column=ci + 1, value=val).border = thin_border
            ri += 1

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    toplam_row_idx = ri + 1
    ws.cell(row=toplam_row_idx, column=1, value="TOPLAM").font = Font(bold=True)
    for ci in range(1, len(BASIT_USUL_KOLONLAR) + 1):
        ws.cell(row=toplam_row_idx, column=ci).fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        ws.cell(row=toplam_row_idx, column=ci).font = Font(bold=True)
        ws.cell(row=toplam_row_idx, column=ci).border = thin_border

    brut_kolon = BASIT_USUL_KOLONLAR.index("GENEL TOPLAM") + 1 if "GENEL TOPLAM" in BASIT_USUL_KOLONLAR else 29
    toplam_brut = sum((r.get("brut", 0) or 0) for r in results if "error" not in r)
    ws.cell(row=toplam_row_idx, column=brut_kolon, value=toplam_brut)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
